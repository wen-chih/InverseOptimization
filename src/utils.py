# src/utils.py

import numpy as np
import gurobipy as gp
from gurobipy import GRB

def solve_LP(p, A, b, A_eq=None, b_eq=None, x_ub=None):
    """
    Forward Problem Solver
    Max p^T x
    s.t. Ax <= b
         A_eq x = b_eq (Optional)
         0 <= x <= x_ub (Optional)
    """

    if p is None:
        return None

    n = len(p)
    
    # Create the model
    model = gp.Model("ForwardProblem")
    model.setParam("OutputFlag", 0) 
    
    try:
        # define variables
        ub_val = x_ub if x_ub is not None else GRB.INFINITY
        x = model.addMVar(n, lb=0.0, ub=ub_val, name="x")
        
        # define constraints (Ax <= b)
        model.addConstr(A @ x <= b, name="ineq")
        
        # define constraints (Optional)
        if A_eq is not None:
            model.addConstr(A_eq @ x == b_eq, name="eq")
            
        # Seting objective function (Maximization)
        model.setObjective(p @ x, GRB.MAXIMIZE)
        
        # solving the model
        model.optimize()
        
        if model.status == GRB.OPTIMAL:

            return x.X # return optimal x    
        elif model.status == GRB.INFEASIBLE:
            print("No feasible solution exists. Infeasible")  
        elif model.status == GRB.UNBOUNDED:
            print("No feasible solution exists. Unbounded")
        else:
            print("No solution found")
        return None
            
    except Exception as e:
        return None
    finally:
        model.dispose()


def sheet_styling(output_file):

    # ---------------------------------------------------    
    # Open with openpyxl for styling
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(output_file)

    font = Font(name="Calibri", size=12)   # or "Calibri", "Arial", etc.
    for ws in wb.worksheets:            # loop through all sheets
        for row in ws.iter_rows():      # loop through all rows
            for cell in row:            # loop through cells
                cell.font = font

    wb.save(output_file)

  